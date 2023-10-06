'''
sokolov: An LLM-experimenting suite for my singular 'they' research. Will include data migration, cleanup, and checking features.

sokolov sokolov sokolov!
sokolov?
That's right.
'''


# standard libraries
import os
import re
import logging

# installed libraries
import openpyxl
from openpyxl.styles import PatternFill, Font, NamedStyle


# project resources
from argparse_assets import handle_args
from openai_assets import moderation_check, conduct_experiment, count_tokens
from excel_assets import *


def translate_annotation(they_type: str) -> str:
    '''Translate the annotation codes into either of the three major categories.'''
    if they_type in ['plural_they', 'collnoun']:
        return "plural"
    else:
        return "singular"


def get_ord(word: str):
	return num2words(word, ordinal=True)

def read_span(text: str) -> tuple:
    return (int(text.split(',')[0][1:]), int(text.split(',')[1][1:-1]))

def adapted(prompt: str, body: str, span: str) -> str:
    span = read_span(span)
    they_form = body[span[0]:span[1]]
    if len(re.findall(they_form.lower(), body.lower())) == 1:
        ordinal = ''
    else:
        matches = [elem.span()[0] for elem in re.finditer(they_form.lower(), body.lower())]
        position = matches.index(span[0]) + 1
        ordinal = f' {get_ord(position)}'

    return prompt.format(ordinal, they_form, body)


def process_experiment_file(file: str, args):
    # Load workbook and sheet
    wb = openpyxl.load_workbook(file)
    data_sheet = wb['data']
    last_row = get_last_row_with_data(data_sheet, "A")

    # Remove unannotated comments
    rows_to_delete = []
    for idx, row in enumerate(data_sheet.iter_rows(min_row=2, max_row=last_row, values_only=True), 2):
        annotated_col_idx = [cell.value for cell in data_sheet[1]].index("annotated")
        if row[annotated_col_idx] != "X":
            rows_to_delete.append(idx)

        noise_col_idx = [cell.value for cell in data_sheet[1]].index("noise")
        if row[noise_col_idx] == "X":
            rows_to_delete.append(idx)

        # Remove comments based on filter_length
        if args.length:
            comment_body_col_idx = [cell.value for cell in data_sheet[1]].index("comment_body")
            body = row[comment_body_col_idx]
            span_col_idx = [cell.value for cell in data_sheet[1]].index("span")
            span = row[span_col_idx]
            if args.unit == 'tokens':
                tk_length = count_tokens(adapted(args.prompt, body, span))
                if tk_length > args.length:
                    rows_to_delete.append(idx)
            elif args.unit == 'chars':
                if body > args.length:
                    rows_to_delete.append(idx)

    # delete all rows marked for deletion
    for idx in reversed(rows_to_delete):
        data_sheet.delete_rows(idx)

    last_row = get_last_row_with_data(data_sheet, "A")

    # Add new columns
    last_col = data_sheet.max_column
    headers = ["prompt", "human_annotation", "LLM_response", "LLM_annotation", "inter-annotator_agreement"]
    for idx, header in enumerate(headers, start=1):
        data_sheet.cell(row=1, column=last_col + idx, value=header)

    # Fill in 'prompt' and 'human_annotation' column
    for idx, row in enumerate(data_sheet.iter_rows(min_row=2, max_row=last_row, values_only=True), 2):
        comment_body_col_idx = [cell.value for cell in data_sheet[1]].index("comment_body")
        body = row[comment_body_col_idx]

        human_col_idx = [cell.value for cell in data_sheet[1]].index("human_annotation")

        ambig_col_idx = [cell.value for cell in data_sheet[1]].index("ambiguous")
        ambig = row[ambig_col_idx]

        they_type_idx = [cell.value for cell in data_sheet[1]].index("they_type")
        they_type = row[they_type_idx]

        span_col_idx = [cell.value for cell in data_sheet[1]].index("span")
        span = row[span_col_idx]

        prompt_col_idx = [cell.value for cell in data_sheet[1]].index("prompt")

        if data_sheet.cell(row=idx, column=prompt_col_idx+1).value is None:
            data_sheet.cell(row=idx, column=prompt_col_idx+1).value = adapted(args.prompt, body, span)
        
        if data_sheet.cell(row=idx, column=human_col_idx+1).value is None:
            data_sheet.cell(row=idx, column=human_col_idx+1).value = "ambiguous" if ambig == "X" else translate_annotation(they_type)

    # Add results sheet
    if 'results' not in wb.sheetnames:
        results_sheet = wb.create_sheet('results')
    else:
        results_sheet = wb['results']

    categories = ["singular they", "plural they", "ambiguous they"]
    metrics = ["True Positives", "False Positives", "True Negatives", "False Negatives", "Recall", "Precision", "Accuracy"]

    # Headers
    for idx, metric in enumerate(metrics, start=2):
        results_sheet.cell(row=1, column=idx, value=metric)
        if metric in ["Recall", "Precision", "Accuracy"]:
            results_sheet.cell(row=1, column=idx).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            results_sheet.cell(row=1, column=idx).font = Font(bold=True)
    # Rows
    for idx, category in enumerate(categories, start=2):
        results_sheet.cell(row=idx, column=1, value=category)

    directory = os.path.dirname(file)

    if args.moderation:
        filtered_file = os.path.join(directory, "filtered-out-comments.xlsx")
        moderation_check(data_sheet, filtered_file)

    # Metrics calculation
    data_sheet_max_row = last_row
    human_annotation_col_letter = get_column_letter([cell.value for cell in data_sheet[1]].index("human_annotation") + 1)
    llm_annotation_col_letter = get_column_letter([cell.value for cell in data_sheet[1]].index("LLM_annotation") + 1)
    wb.add_named_style(NamedStyle(name="percent_style", number_format="0.0%"))
    add_metric_formulas(results_sheet, 'data', human_annotation_col_letter, llm_annotation_col_letter, data_sheet_max_row)

    add_statistics(wb)

    # Save file
    directory = os.path.dirname(file)
    new_filepath = os.path.join(directory, "experiment.xlsx")
    wb.save(new_filepath)


def main():
    #logging.basicConfig(level=logging.NOTSET, format='INFO: %(message)s')
    args = handle_args()

    file = args.inputfile

    if args.task == "preparation":
        process_experiment_file(file, args)
    elif args.task == "experiment":
        conduct_experiment(file, args.llm)


if __name__ == "__main__":
    main()