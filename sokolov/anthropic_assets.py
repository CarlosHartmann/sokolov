'''
anthropic_assets: All things related to the Anthropic API.
NOTE: Moderation check not yet implemented for anthropic!
'''

import spacy
nlp = spacy.load("en_core_web_sm")

import openpyxl
import requests
import time

from sokolov.excel_assets import *

import anthropic
client = anthropic.Anthropic(
    api_key="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",
)


import tiktoken
encoding = tiktoken.get_encoding("cl100k_base")



def read_span(text: str) -> tuple:
    return (int(text.split(',')[0][1:]), int(text.split(',')[1][1:-1]))


def count_tokens(text: str) -> int:
    """Counts the length of input text in GPT tokens."""
    num_tokens = len(encoding.encode(text))
    return num_tokens


def get_llm_response(prompt, llm):
    message = client.messages.create(
    model=llm,
    max_tokens=512,
    messages=[
        {"role": "user", "content": prompt}
    ])
    
    return message.content[0].text


def complete_statistics(results_sheet, they_type_col, outside_col, unknow_col, IAA_col, ambiguous_col):
    plural_col = 12
    singular_col = plural_col+3

    for col in [plural_col, singular_col]:
        letter = get_column_letter(col)
        lastrow = get_last_row_with_data(results_sheet, letter)
        for row in range(2, lastrow+1):
            category = results_sheet.cell(row=row, column=col).value
            cat, extension = category.split()[0], category.split()[1] if len(category.split()) > 1 else None
            if not extension:
                iaa_num = f'COUNTIFS(data!{they_type_col}:{they_type_col}, "{cat}", data!{outside_col}:{outside_col}, "<>X", data!{IAA_col}:{IAA_col}, "X")'
            elif extension == "unknowable":
                iaa_num = f'COUNTIFS(data!{they_type_col}:{they_type_col}, "{cat}", data!{unknow_col}:{unknow_col}, "X", data!{IAA_col}:{IAA_col}, "X")'
            elif extension == "outside":
                iaa_num = f'COUNTIFS(data!{they_type_col}:{they_type_col}, "{cat}", data!{unknow_col}:{unknow_col}, "<>X", data!{outside_col}:{outside_col}, "X", data!{IAA_col}:{IAA_col}, "X")'
            
            comparison_col = get_column_letter(col+1)
            calculation = f'({iaa_num})/{comparison_col}{row}'
            formula = f'=IF({comparison_col}{row}>0,{calculation},"--")'

            results_sheet.cell(row=row, column=col+2).value = formula
            results_sheet.cell(row=row, column=col+2).number_format = '0.00%'
    
    lastrow += 1
    ambig_iaa_col = singular_col+5
    comparison_col_letter = get_column_letter(ambig_iaa_col-1)

    iaa_num = f'COUNTIFS(data!{ambiguous_col}:{ambiguous_col}, "X", data!{IAA_col}:{IAA_col}, "X")'
    calculation = f'({iaa_num}/{comparison_col_letter}{lastrow})'
    results_sheet.cell(row=lastrow, column=ambig_iaa_col).value = f'=IF({comparison_col_letter}{lastrow}>0,{calculation}, "--")'
    results_sheet.cell(row=lastrow, column=ambig_iaa_col).number_format = '0.00%'

    # total IAA %
    for col in [plural_col, singular_col]:
        iaa_col = col+2
        iaa_letter = get_column_letter(iaa_col)
        amt_letter = get_column_letter(iaa_col-1)
        last_data_row = get_last_row_with_data(results_sheet, column=iaa_letter)
        formula = f'=SUMPRODUCT({amt_letter}2:{amt_letter}{last_data_row}, {iaa_letter}2:{iaa_letter}{last_data_row}) / SUM({amt_letter}2:{amt_letter}{last_data_row})'
        results_sheet.cell(row=lastrow, column=iaa_col).value = formula



def is_obvious_case(body, span):
    span = read_span(span)
    body = body.lower()  # Convert the text to lowercase to ensure case insensitivity
    start, end = span[0], span[1]

    doc = nlp(body)
    
    # Find the spaCy token corresponding to the start index
    start_token_index = None
    for token in doc:
        if token.idx == start:
            start_token_index = token.i
            break

    # Handle cases where the start token is not found
    if start_token_index is None:
        return False
    
    # Extract the pronoun
    pronoun = doc[start_token_index].text.lower()

    # Check for "NUM of them"
    if start_token_index > 0 and doc[start_token_index - 2].pos_ == "NUM" and doc[start_token_index - 1].text == "of" and pronoun == "them":
        return True
    
    # Check for "they VERB each other" or "they VERB one another"
    if pronoun == "they" and start_token_index < len(doc) - 3:
        following_phrase = ' '.join([doc[start_token_index + 2].text.lower(), doc[start_token_index + 3].text.lower()])
        if doc[start_token_index + 1].pos_ == "VERB" and following_phrase in ["each other", "one another"]:
            return True

    # Get last three words before the pronoun and first three words after the pronoun
    preceding_text = body[:start].strip().split()[-3:]  
    following_text = body[end:].strip().split()[:3]

    # Joining the words to form phrases to match with the expressions
    last_two_words = ' '.join(preceding_text[-2:])
    last_three_words = ' '.join(preceding_text)
    next_word = ' '.join(following_text[:1])
    next_two_words = ' '.join(following_text[:2])
    next_three_words = ' '.join(following_text)

    # List of special expressions
    special_expressions = [
        "all of them",
        "some of them",
        "most of them",
        "both of them",
        "pair of them",
        "either of them",
        "they are all",
        "they were all",
        "they will all",
        "they all"
    ]

    # Checking if the phrase is in the list of special expressions
    is_special_expression = ' '.join([last_two_words, pronoun]) in special_expressions or \
                            ' '.join([last_three_words, pronoun]) in special_expressions

    # Checking phrases that include words after the pronoun
    is_special_expression = is_special_expression or \
                            ' '.join([pronoun, next_word]) in special_expressions or \
                            ' '.join([pronoun, next_two_words]) in special_expressions or \
                            ' '.join([pronoun, next_three_words]) in special_expressions
                            
    return is_special_expression


def clauduct_experiment(file: str, llm: str):
    '''
    Central function for conducting experiments.
    Requires a workbook that will have the necessary headers for the experiment (prompt, LLM_response, LLM_annotation, human_annotation, inter-annotator-agreement).
    Also requires a "results" worksheet that will be extended with the total IAA and should already have the necessary formulas to display the overall results.
    '''

    # initiate two required worksheets
    workbook = openpyxl.load_workbook(file)
    data_sheet = workbook['data']
    results_sheet = workbook['results']

    # get required column numbers
    # for the data sheet
    prompt_col = get_column_by_header(data_sheet, 'prompt', out="num")
    response_col = get_column_by_header(data_sheet, 'LLM_response', out="num")
    response_letter = get_column_by_header(data_sheet, 'LLM_response', out="letter")
    annotation_col = get_column_by_header(data_sheet, 'LLM_annotation', out="num")
    annotation_letter = get_column_by_header(data_sheet, 'LLM_annotation', out="letter")
    human_annotation_col = get_column_by_header(data_sheet, 'human_annotation', out="num")
    human_annotation_letter = get_column_by_header(data_sheet, 'human_annotation', out="letter")
    IAA_col = get_column_by_header(data_sheet, 'inter-annotator_agreement', out="num")

    # for the plural_they filter
    body_col = get_column_by_header(data_sheet, 'comment_body', out="num")
    span_col = get_column_by_header(data_sheet, 'span', out="num")

    # for statistics
    they_type_col = get_column_by_header(data_sheet, 'they_type')
    outside_col = get_column_by_header(data_sheet, 'referent_outside')
    unknow_col = get_column_by_header(data_sheet, 'unknowable')
    ambiguous_col = get_column_by_header(data_sheet, 'ambiguous')

    lastrow = get_last_row_with_data(data_sheet)

    thread = None # for potential use of assistants

    for row in range(2, lastrow+1):
        if row % 20 == 0:
            workbook.save(file)
        prompt = data_sheet.cell(row=row, column=prompt_col).value # read prompt

        #for plural_they filter
        body = data_sheet.cell(row=row, column=body_col).value
        span = data_sheet.cell(row=row, column=span_col).value

        if not data_sheet.cell(row=row, column=response_col).value: # don't wanna redo what's already been requested before
            if not is_obvious_case(body, span):          
                response = get_llm_response(prompt, llm) # send it to LLM

                try:    
                    data_sheet.cell(row=row, column=response_col).value = response
                except ValueError:
                    print(f"The following response could not be pasted into excel. It was supposed to go in row {row}:")
                    print(response)
            else:
                data_sheet.cell(row=row, column=response_col).value = "Per rule-based 'plural they' filter it is plural."

        data_sheet.cell(row=row, column=annotation_col).value = f'''=IF(ISNUMBER(SEARCH("singular", {response_letter}{row})), "singular", IF(ISNUMBER(SEARCH("plural", {response_letter}{row})), "plural", IF(ISNUMBER(SEARCH("ambiguous", {response_letter}{row})), "ambiguous", "")))'''

        llm_annotation = data_sheet.cell(row=row, column=annotation_col).value
        human_annotation = data_sheet.cell(row=row, column=human_annotation_col).value
        data_sheet.cell(row=row, column=IAA_col).value = f"=IF({human_annotation_letter}{row} = {annotation_letter}{row}, \"X\", \"\")"
    
    workbook.save(file) # save before doing the statistics part

    # Metrics calculation
    human_annotation_col_letter = get_column_letter([cell.value for cell in data_sheet[1]].index("human_annotation") + 1)
    llm_annotation_col_letter = get_column_letter([cell.value for cell in data_sheet[1]].index("LLM_annotation") + 1)
    add_metric_formulas(results_sheet, 'data', human_annotation_col_letter, llm_annotation_col_letter, lastrow)

    complete_statistics(results_sheet, they_type_col, outside_col, unknow_col, get_column_letter(IAA_col), ambiguous_col)

    workbook.save(file)