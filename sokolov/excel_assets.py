'''
excel_assets: Everything needed to handle and manipulate excel files in this project.
'''


import re
from openpyxl.utils import get_column_letter
from num2words import num2words


def add_metric_formulas(results_sheet, data_sheet, human_col_letter, llm_col_letter, data_sheet_max_row):
    categories = ["singular", "plural", "ambiguous"]
    
    for idx, category in enumerate(categories, start=2):
        tp_formula = f'SUMPRODUCT(({data_sheet}!{human_col_letter}$2:{human_col_letter}${data_sheet_max_row}="{category}")*({data_sheet}!{llm_col_letter}$2:{llm_col_letter}${data_sheet_max_row}="{category}"))'
        fp_formula = f'SUMPRODUCT(({data_sheet}!{human_col_letter}$2:{human_col_letter}${data_sheet_max_row}<>"{category}")*({data_sheet}!{llm_col_letter}$2:{llm_col_letter}${data_sheet_max_row}="{category}"))'
        tn_formula = f'SUMPRODUCT(({data_sheet}!{human_col_letter}$2:{human_col_letter}${data_sheet_max_row}<>"{category}")*({data_sheet}!{llm_col_letter}$2:{llm_col_letter}${data_sheet_max_row}<>"{category}"))'
        fn_formula = f'SUMPRODUCT(({data_sheet}!{human_col_letter}$2:{human_col_letter}${data_sheet_max_row}="{category}")*({data_sheet}!{llm_col_letter}$2:{llm_col_letter}${data_sheet_max_row}<>"{category}"))'
        
        results_sheet.cell(row=idx, column=2).value = f'={tp_formula}'
        results_sheet.cell(row=idx, column=3).value = f'={fp_formula}'
        results_sheet.cell(row=idx, column=4).value = f'={tn_formula}'
        results_sheet.cell(row=idx, column=5).value = f'={fn_formula}'
        
        recall_formula = f'={get_column_letter(2)}{idx}/({get_column_letter(2)}{idx}+{get_column_letter(5)}{idx})'
        precision_formula = f'={get_column_letter(2)}{idx}/({get_column_letter(2)}{idx}+{get_column_letter(3)}{idx})'
        accuracy_formula = f'({get_column_letter(2)}{idx}+{get_column_letter(4)}{idx})/({get_column_letter(2)}{idx}+{get_column_letter(3)}{idx}+{get_column_letter(4)}{idx}+{get_column_letter(5)}{idx})'
        
        results_sheet.cell(row=idx, column=6).value = f'={recall_formula}'
        results_sheet.cell(row=idx, column=7).value = f'={precision_formula}'
        results_sheet.cell(row=idx, column=8).value = f'={accuracy_formula}'
    
    for row in results_sheet.iter_rows(min_row=2, max_row=4, min_col=6, max_col=8):
        for cell in row:
            cell.style = "percent_style"


def get_last_row_with_data(sheet, column="A"):
    # Loop from the bottom up until a non-empty cell is found
    for row in range(sheet.max_row, 0, -1):
        if sheet[f"{column}{row}"].value:
            return row
    return None  # If no data found


def get_column_by_header(worksheet, header_name):
    """Returns the column letter of a given header_name in worksheet."""
    for column in worksheet.iter_cols(1, worksheet.max_column):
        if column[0].value == header_name:
            return get_column_letter(column[0].column)
    return None


def add_statistics(workbook):
    results_sheet = workbook['results']
    data_sheet = workbook['data']
    # Adding headers
    headers = ['plural', 'singular', 'ambiguous']
    categories = {'plural': ['plural_they', 'plural_they outside', 'plural_they unknowable', 'collnoun', 'collnoun outside', 'collnoun unknowable'],
                  'singular': ['mixed_they', 'mixed_they outside', 'mixed_they unknowable', 'generic_singularthey', 'generic_singularthey outside', 'generic_singularthey unknowable', 'generic_genderedthey', 'generic_genderedthey outside', 'generic_genderedthey unknowable', 'specific_unknown_singularthey', 'specific_unknown_singularthey outside', 'specific_unknown_singularthey unknowable', 'specific_singularthey', 'specific_singularthey outside', 'specific_singularthey unknowable', 'nonbinary_they', 'nonbinary_they outside', 'nonbinary_they unknowable']}
    col = 12
    for header in headers:
        results_sheet.cell(row=1, column=col, value=header)
        results_sheet.cell(row=1, column=col+1, value='amount')
        results_sheet.cell(row=1, column=col+2, value='IAA')
        col += 3

    for number in list(categories.keys()):
        col = 13 if number == "plural" else 16
        row = 1 # incremented before each writing-in step
        for category in categories[number]:
            row += 1

            # category name
            results_sheet.cell(row=row, column=col-1).value = category

            # Get column letters for the required headers in data worksheet
            they_type_col = get_column_by_header(data_sheet, "they_type")
            referent_outside_col = get_column_by_header(data_sheet, "referent_outside")
            unknowable_col = get_column_by_header(data_sheet, "unknowable")
            ambiguous_col = get_column_by_header(data_sheet, "ambiguous")

            category_simple = category.split()[0]

            # amount of instances in data
            if 'outside' in category:
                results_sheet.cell(row=row, column=col).value = f'=COUNTIFS(data!{they_type_col}:{they_type_col}, "{category_simple}", data!{referent_outside_col}:{referent_outside_col}, "X", data!{unknowable_col}:{unknowable_col}, "<>X")'
            elif 'unknowable' in category:
                results_sheet.cell(row=row, column=col).value = f'=COUNTIFS(data!{they_type_col}:{they_type_col}, "{category_simple}", data!{referent_outside_col}:{referent_outside_col}, "X", data!{unknowable_col}:{unknowable_col}, "X")'
            else:
                results_sheet.cell(row=row, column=col).value = f'=COUNTIFS(data!{they_type_col}:{they_type_col}, "{category_simple}", data!{referent_outside_col}:{referent_outside_col}, "<>X")'
        
        # insert sum formula for "total" line
        lowest_row = len(categories['singular']) + 1 # singular will always have the most categories, so the lowest row with data will be the header-row + singular rows
        col_letter = get_column_letter(col)
        results_sheet.cell(row=lowest_row+1, column=col).value = f'=SUM({col_letter}2:{col_letter}{row})'

    # insert total number of ambiguous cases in the data
    col += 3
    row = lowest_row+1
    results_sheet.cell(row=row, column=col).value = f'=COUNTIFS(data!{ambiguous_col}:{ambiguous_col}, "X")'

    # insert total number of data points
    col += 2
    results_sheet.cell(row=row, column=col).value = f'=SUM(M{row},P{row},S{row})'  

    results_sheet.cell(row=20, column=11).value = "Total"

    # Find the max length from plural and singular lists to determine the "Total" row position
    max_len = max(len(categories['plural']), len(categories['singular']))
    results_sheet[f"K{2+max_len}"] = "Total"

    # Auto size columns based on content starting from column K
    for col in ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]:
        results_sheet.column_dimensions[col].bestFit = True